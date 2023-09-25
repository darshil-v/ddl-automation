from openpyxl import load_workbook
import os

filename = "SAMPLE_TBL-STRUCTURE.xlsx"
wb = load_workbook(filename)
print(f"File: {filename}\n")

datatype_file = "sf_datatypes.txt"
with open(datatype_file) as f:
    dtypes = [line.rstrip() for line in f]


def checkDatatype(idx, row, headers, errcount=0):
    if row[2].upper() not in dtypes:
        print(f"{sheet} (C{idx+2}): Invalid data entry for {headers[2]}: {row[2]}")
        errcount += 1

    return errcount


def checkLenScale(idx, row, headers, errcount=0):

    if (
        (not isinstance(row[3], int))
        and (row[2][:4].upper() != "DATE")
        and (row[2][:4].upper() != "TIME")
    ):
        print(
            f"{sheet} (D{idx+2}): Invalid data entry for {row[2]}: {headers[3]} {row[3]}"
        )
        errcount += 1

    if (
        (row[2].upper() in dtypes[:15])
        and (not isinstance(row[4], int))
        and (row[2].upper() != "DATE")
    ):
        print(
            f"{sheet} (E{idx+2}): Invalid data entry for {row[2]}: {headers[4]} {row[4]}"
        )
        errcount += 1

    return errcount


def checkNull(idx, row, headers, errcount=0):
    if not isinstance(row[5], bool):
        print(f'{sheet} (F{idx+2}): Invalid data entry for {headers[5]}: "{row[5]}"')
        errcount += 1

    return errcount


def checkConstraints(idx, row, headers, errcount=0):
    if row[6]:
        if row[6][0] not in "PpUuFf":
            print(
                f'{sheet} (G{idx+2}): Invalid {headers[6]} - "{row[6]}". \nPlease check before proceeding.'
            )
            errcount += 1
        elif row[6][0] in "Ff":
            if not row[7] or not row[8]:
                print(
                    f"{sheet} (G{idx+2}): Reference fields missing for foreign key(s)"
                )
                errcount += 1

    return errcount


for sheet in wb.sheetnames:
    headers = []
    for i in wb[sheet][1]:
        headers.append(i.value)
    count = 0
    for idx, row in enumerate(wb[sheet].iter_rows(min_row=2, values_only=True)):
        count += checkDatatype(idx, row, headers)
        count += checkLenScale(idx, row, headers)
        count += checkNull(idx, row, headers)
        count += checkConstraints(idx, row, headers)

    print(f"Total errors in {sheet}: {count}\n")

wb.close()
