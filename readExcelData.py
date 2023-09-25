import logging
import sys
import os
from openpyxl import load_workbook
from SFConnectionParser import SFConnection
import verifyLocalData as vld


class ReadExcelData:
    def __init__(self):
        self.log_level = logging.INFO
        self.log_fmt = "[%(levelname)s] - %(message)s"
        logging.basicConfig(format=self.log_fmt)

    def constraintKeyCheck(self, ck):
        if ck[0] in "Pp":
            return "PRIMARY KEY"
        elif ck[0] in "Uu":
            return "UNIQUE"

    def extractExcelFooterData(self, wb, sheet):
        dtemp = {"PRIMARY KEY": [], "UNIQUE": []}
        stemp = ""
        try:
            for (
                col_num,
                col_name,
                dtype,
                col_len,
                col_scale,
                not_null,
                *constraints,
            ) in wb[sheet].iter_rows(min_row=2, values_only=True):
                if constraints != [None]:
                    dtemp[self.constraintKeyCheck(constraints[0])].append(col_name)
            for k, v in dtemp.items():
                if not v:
                    dtemp.pop(k)
                    break
                stemp += f"{k} ("
                for i in v:
                    stemp += f'"{i}",'
                stemp = stemp[:-1]
                stemp += f"),\n"

            return stemp

        except ValueError as ve:
            logging.info(ve)
        except TypeError as te:
            print(te)

        # headers = list()
        # for i in wb[sheet][1]:
        #    headers.append(i.value)

        ### USEFUL LATER
        ##ckeys = {0:'Primary Key', 1:'Foreign Key', 2:'Unique Key'}

        # if constraint_keys:
        #    ck = self.constraintKeyCheck(constraint_keys)
        #    footer_info += f"{ck} (\"{col_name}\"),\n"
        #### USEFUL LATER
        # temp_footer = [(ckeys[i],col_name) for i,x in enumerate(constraint_keys) if x]

    def extractExcelBodyData(self, wb, sheet):
        # temp statment to process each sheet
        tmpStmt = ""
        stmt = ""

        try:
            # headers = list()
            # for i in wb[sheet][1]:
            #    headers.append(i.value)
            for (
                col_num,
                col_name,
                dtype,
                col_len,
                col_scale,
                not_null,
                *constraints,
            ) in wb[sheet].iter_rows(min_row=2, values_only=True):
                # check for col_name, dtype, precision/scale, null_cond
                if (isinstance(col_len, int)) and (
                    col_scale == "[NULL]" or not col_scale
                ):
                    dtype = f"{dtype}({col_len})"
                elif (col_len == "[NULL]" or not col_len) and (
                    isinstance(col_scale, int)
                ):
                    dtype = f"{dtype}({col_scale})"
                elif (isinstance(col_len, int)) and (isinstance(col_scale, int)):
                    dtype = f"{dtype}({col_len},{col_scale})"

                if not_null:
                    null_cond = "NOT NULL"
                else:
                    null_cond = "NULL"

                if col_name and dtype:
                    # TODO: Add error logging
                    tmpStmt = f'"{col_name}" {dtype} {null_cond}'
                    tmpStmt = tmpStmt + ",\n"
                # TODO: format into try-finally block after logging
                stmt = stmt + tmpStmt

            return stmt

        except ValueError as ve:
            logging.info(
                ve
            )  # (f"WARNING: ValueError: Missing constraints in table: {sheet}")
            # footer_info = "\n"+");\n"

    def writeSQLFile(
        self,
        sheet,
        content,
        current_db,
        current_schema,
        current_wh,
        current_role,
        dirPathSrc,
    ):
        try:
            # logging.info('Writing sql to file', sheet)
            # open sql file located in appropriate sub-directory
            filename = f"{current_db}.{current_schema}.{sheet}.sql"
            with open(os.path.join(dirPathSrc, filename), "w+") as f:
                f.write(f"USE DATABASE {current_db};\n")
                f.write(f"USE SCHEMA {current_schema};\n")
                f.write(f"USE WAREHOUSE {current_wh};\n")
                f.write(f"USE ROLE {current_role};\n")
                f.write("\n")
                f.write(f"{content}")
                f.write(");\n")

        except FileNotFoundError:
            # If error is not already exists, then raise the error else continue
            return f"File not found!"

    def readExcelWorkbook(self, fileName, dirPathSrc, user="snowflake"):
        try:
            # load workbook using openpyxl
            wb = load_workbook(fileName)
            # f1 = open("excel_tables.txt","w+")

            # query role, warehouse, db, schema
            cfg = vld.readConfigFile(vld.getDirPath())
            # res_metadata = cur.execute("select current_role(), current_warehouse(), current_database(), current_schema();").fetchall()
            current_role = cfg[user]["role"]
            current_wh = cfg[user]["warehouse"]
            current_db = cfg[user]["database"]
            current_schema = cfg[user]["schema"]
            replace_flag = cfg["local_files"]["replace_sql"]

            # process individual sheet (table name) in workbook
            for sheet in wb.sheetnames:
                print(f"Processing {sheet}...")
                ### DDL Header <CREATE TABLE>
                if replace_flag == "True":
                    header_detail = f"CREATE OR REPLACE TABLE {current_db}.{current_schema}.{sheet}(\n"
                else:
                    header_detail = f"CREATE TABLE IF NOT EXISTS {current_db}.{current_schema}.{sheet}(\n"

                ### DDL Body <coln_name, datatype, null_cond>
                try:
                    # extract sheet data
                    body_detail = self.extractExcelBodyData(wb, sheet)
                except TypeError as te:
                    logging.info(
                        f"WARNING: Table {sheet} missing required fields! Table skipped!\n"
                    )
                    continue
                except Exception as e:
                    print(e)
                    continue

                ### DDL Footer <constraint keys>

                footer_detail = self.extractExcelFooterData(wb, sheet)

                # if sheet extracted, combine query contents
                if body_detail:
                    content = header_detail + body_detail + footer_detail
                else:
                    print(f"{sheet} missing data..")
                    # print(content[:-2])
                # write to sql file
                self.writeSQLFile(
                    sheet,
                    content[:-2],
                    current_db,
                    current_schema,
                    current_wh,
                    current_role,
                    dirPathSrc,
                )

        except ModuleNotFoundError as mdfe:
            logging.error("Required package missing!")
            sys.exit(1)

    """
    def readExcelWorkbookSheet(self, fileName, sheetList, dirPathSrc, cur, user = 'snowflake'):
        try:
            #load workbook using openpyxl
            wb = load_workbook(fileName)
            #f1 = open("excel_tables.txt","w+")

            #query role, warehouse, db, schema
            cfg = vld.readConfigFile(vld.getDirPath())
            #res_metadata = cur.execute("select current_role(), current_warehouse(), current_database(), current_schema();").fetchall()
            current_role =  cfg[user]['role']
            current_wh = cfg[user]['warehouse']
            current_db = cfg[user]['database']
            current_schema = cfg[user]['schema']
            #= res_metadata[0][0], res_metadata[0][1], res_metadata[0][2], res_metadata[0][3]

            #process individual sheet (table name) in workbook
            for sheet in sheetList:
                sheet = sheet.split('.')[-1]
                header_detail = f"CREATE TABLE IF NOT EXISTS {current_db}.{current_schema}.{sheet}(\n"
                try:
                    #extract sheet data
                    body_detail, footer_detail = self.extractExcelBodyData(wb,sheet)
                except TypeError as te:
                    logging.info(f"WARNING: Table {sheet} missing required fields! Table skipped!\n")
                    logging.info(te)
                    continue
                except Exception as e:
                    logging.error(e)
                #if sheet extracted, combine query contents
                if(body_detail and footer_detail):
                    content = header_detail + body_detail + footer_detail
                    #print(content)
                #create sub-directory if not exists
                #if not os.path.exists(os.path.join(dirPathSrc, sheet)):
                #    os.makedirs(os.path.join(dirPathSrc, sheet))
                #    print("Created new directory:", sheet)

                #write to sql file
                self.writeSQLFile(sheet, content, current_db, current_schema, current_wh, current_role, dirPathSrc)
                #f1.write(current_db+"."+current_schema+"."+sheet+"\n")
            #f1.close()

        except ModuleNotFoundError as mdfe:
            logging.error("Required package missing!")
            sys.exit(1)
    """
